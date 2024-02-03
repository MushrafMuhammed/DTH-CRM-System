from email import message
from django.contrib.auth import authenticate, login, logout
from django.http import JsonResponse
from django.shortcuts import redirect, render
from openpyxl import load_workbook
from .forms import LeadAssignForm
from .models import Administrator, Lead, Telecaller
from django.core.mail import send_mail
from dth_crm import settings


# Create your views here.

def loginfun(request):
    # if request.user.is_authenticated:
    #     return redirect('adminapp:dashboard')
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        # user = authenticate(username=username, password=password)
        try :
            admin = Administrator.objects.get(
                 username = username,
                 password = password
            )
            request.session["admin_sessionID"] = admin.id
            return redirect('adminapp:dashboard')
        except :
            print('invalid username or password !')

        # if user is not None:
        #     login(request, user)
        #     return redirect('adminapp:dashboard')
        # else:
        #     print('Invalid credentials') 
    return render(request, 'adminapp/pages-login.html')

def registerfun(request):
    return render(request, 'adminapp/pages-register.html')

def layoutfun(request):
    administrator = request.session["admin_sessionID"]
    profile = Administrator.objects.get(id=administrator)
    return render(request, 'adminapp/components-layout.html',{'profile' : profile.name})

def dashboardfun(request):
    if 'admin_sessionID' in request.session:
        return render(request, 'adminapp/pages-dashboard.html')
    else:
        return redirect('adminapp:login')

def customerfun(request):
    return render(request, 'adminapp/pages-customer.html')

def peoplefun(request):
    return render(request, 'adminapp/pages-people.html')

def companyfun(request):
    return render(request, 'adminapp/pages-company.html')

def leadfun(request):
    if 'admin_sessionID' in request.session:
        leads = Lead.objects.all()

        if request.method == 'POST':
            print("***************tested")
            excel_file = request.FILES['excel_file']
            wb = load_workbook(excel_file)
            ws = wb.active

            existing_phones = set(Lead.objects.values_list('phone', flat=True))

            for row in ws.iter_rows(min_row=2, values_only=True):
                type, lead_status, source, country, name, phone, email, status = row
                
                # Check if the phone numbers already exists
                if phone in existing_phones:
                    error_message = 'Phone number already exists in the Lead table.'
                    return render(request, 'adminapp/pages-lead.html', {'error_message': error_message, 'leads': leads })
                
                # If not, create a new Lead
                Lead.objects.create(type=type, lead_status=lead_status, source=source, country=country, name=name, phone=phone, email=email, status=status)
            return render(request, 'adminapp/pages-lead.html', {'success_message': 'Leads imported successfully.', 'leads': leads })
        
        return render(request, 'adminapp/pages-lead.html', {'leads': leads})
    else:
        return redirect('adminapp:login')
def add_leadfun(request):
    try:
        if request.method == 'POST' and request.POST.get('id') == '':
            id = request.POST.get('id')
            type = request.POST.get('type')
            name = request.POST.get('name')
            lead_status = request.POST.get('lead_status')
            source = request.POST.get('source')
            country = request.POST.get('country')
            phone = request.POST.get('phone')
            email = request.POST.get('email')
            print(id)

            Lead.objects.create(
                type=type,
                lead_status=lead_status,
                source=source,
                country=country,
                name=name,
                phone=phone,
                email=email
            )
            return JsonResponse({'message': 'Lead added successfully'})
        
        else:
            # print('update checked')
            lead_id = request.POST.get('id')
            editLeed  = Lead.objects.get(
                id=lead_id
            )

            editLeed.type = request.POST.get('type')
            editLeed.lead_status = request.POST.get('lead_status')
            editLeed.source = request.POST.get('source')
            editLeed.country = request.POST.get('country')
            editLeed.name = request.POST.get('name')
            editLeed.phone = request.POST.get('phone')
            editLeed.email = request.POST.get('email')
            editLeed.save()
            return JsonResponse({'message': 'Lead Updated successfully'})

    except Exception as e:
        return JsonResponse({'message': f'Error: {str(e)}'})

def lead_assignfun(request):
    assigned_leads = []
    print(assigned_leads)
    telecallers = Telecaller.objects.all()

    if request.method == 'POST':
        teleemail = Telecaller.objects.get(id=1)
        form = LeadAssignForm(request.POST)

        if form.is_valid():
            telecaller = form.cleaned_data['telecaller']
            start = form.cleaned_data['start']
            end = form.cleaned_data['end']

            assigned_leads = Lead.objects.filter(
                status='pending'
            )[start:end]

            for lead in assigned_leads:
                lead.status = 'assigned'
                lead.telecaller = telecaller
                lead.save()
                 
            # Send email   
            send_mail(
                'Message form DTH CRM',
                'Now you have assigned a set of Laads. Please check your console.',
                settings.EMAIL_HOST_USER,
                [teleemail.email],  # Check that teleemail.email is valid
                fail_silently=False
            )
        return render(request, 'adminapp/lead_assign.html', {'success_message':"Successfully assigned the data"})

    else:
        form = LeadAssignForm()

    return render(request, 'adminapp/lead_assign.html', {'form': form, 'assigned_leads': assigned_leads, 'telecallers': telecallers})


def offerfun(request):
    return render(request, 'adminapp/pages-offer.html')


def invoicefun(request):
    return render(request, 'adminapp/pages-invoice.html')

def quotefun(request):
    return render(request, 'adminapp/pages-quote.html')

def paymentfun(request):
    return render(request, 'adminapp/pages-payment.html')

def expensesfun(request):
    return render(request, 'adminapp/pages-expense.html')

def productfun(request):
    return render(request, 'adminapp/pages-product.html')

def category_productfun(request):
    return render(request, 'adminapp/pages-product_category.html')


def logoutfun(request):
    if 'admin_sessionID' in request.session:
        request.session.flush()
    return redirect('adminapp:login')


def testfun(request):
    return render(request, 'adminapp/check-layout.html')




